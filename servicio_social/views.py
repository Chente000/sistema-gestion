# servicio_social/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction
from django.db.models import Q
from datetime import time
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
# Importaciones para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
# from openpyxl.drawing.image import Image as OpenpyxlImage # Ya no es necesario si no hay logos
from openpyxl.utils import get_column_letter
from collections import defaultdict 
from django.utils import timezone 
from datetime import datetime
from io import BytesIO
# Importar tus modelos existentes
from .models import ServicioSocial, EstudianteServicioSocial
# Importa tus modelos de programacion:
from programacion.models import Periodo 

# Importar tus formularios y el formset
from .forms import ServicioSocialForm, EstudianteServicioSocialFormSet
# Importaciones para permisos
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from administrador.permissions import PERMISSIONS 

# Para resolver las rutas de archivos estáticos para los logos (ya no necesario para Excel)
# from django.contrib.staticfiles.finders import find 
import os 
from django.conf import settings 


@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_PROYECTO_SERVICIO_SOCIAL),
    login_url='/no_autorizado/') 
def servicio_list(request):
    servicios_qs = ServicioSocial.objects.all().order_by('-periodo_academico__fecha_inicio')

    if not (request.user.is_superuser or request.user.is_super_admin_rol) and \
    not request.user.has_permission(PERMISSIONS.ADD_PROYECTO_SERVICIO_SOCIAL) and \
    not request.user.has_permission(PERMISSIONS.CHANGE_PROYECTO_SERVICIO_SOCIAL) and \
    not request.user.has_permission(PERMISSIONS.DELETE_PROYECTO_SERVICIO_SOCIAL):
        
        estudiante_ids_in_scope = []
        if request.user.carrera_asignada:
            estudiante_ids_in_scope = EstudianteServicioSocial.objects.filter(
                carrera=request.user.carrera_asignada
            ).values_list('id', flat=True)
        elif request.user.departamento_asignado:
            estudiante_ids_in_scope = EstudianteServicioSocial.objects.filter(
                carrera__departamento=request.user.departamento_asignado
            ).values_list('id', flat=True)
        
        if estudiante_ids_in_scope:
            servicios_qs = servicios_qs.filter(estudiantes_participantes__id__in=estudiante_ids_in_scope).distinct()
        else:
            servicios_qs = ServicioSocial.objects.none() 
            messages.warning(request, "No tiene permisos para ver proyectos de servicio social o no tiene una carrera/departamento asignada.")

    filter_nombre_proyecto = request.GET.get('nombre_proyecto')
    filter_tutor_encargado = request.GET.get('tutor_encargado')
    filter_estado = request.GET.get('estado')
    filter_area_accion = request.GET.get('area_accion')
    filter_periodo_id = request.GET.get('periodo')

    if filter_nombre_proyecto:
        servicios_qs = servicios_qs.filter(nombre_proyecto__icontains=filter_nombre_proyecto)
    if filter_tutor_encargado:
        servicios_qs = servicios_qs.filter(
            Q(tutor_nombres__icontains=filter_tutor_encargado) |
            Q(tutor_apellidos__icontains=filter_tutor_encargado)
        )
    if filter_estado:
        servicios_qs = servicios_qs.filter(estado=filter_estado)
    if filter_area_accion:
        servicios_qs = servicios_qs.filter(area_accion_proyecto=filter_area_accion)
    
    if filter_periodo_id:
        servicios_qs = servicios_qs.filter(periodo_academico__id=filter_periodo_id)
            
    estado_choices = ServicioSocial.ESTADO_CHOICES
    area_accion_choices = ServicioSocial.AREA_ACCION_CHOICES
    periodos = Periodo.objects.all().order_by('-fecha_inicio')

    context = {
        'servicios': servicios_qs,
        'filter_nombre_proyecto': filter_nombre_proyecto,
        'filter_tutor_encargado': filter_tutor_encargado,
        'filter_estado': filter_estado,
        'filter_area_accion': filter_area_accion,
        'filter_periodo_id': filter_periodo_id,
        'estado_choices': estado_choices,
        'area_accion_choices': area_accion_choices,
        'periodos': periodos,
        'can_add_servicio': request.user.has_permission(PERMISSIONS.ADD_PROYECTO_SERVICIO_SOCIAL),
        'can_change_servicio': request.user.has_permission(PERMISSIONS.CHANGE_PROYECTO_SERVICIO_SOCIAL),
        'can_delete_servicio': request.user.has_permission(PERMISSIONS.DELETE_PROYECTO_SERVICIO_SOCIAL),
        'can_view_detail_servicio': request.user.has_permission(PERMISSIONS.VIEW_PROYECTO_SERVICIO_SOCIAL),
        'can_export_pdf': request.user.has_permission(PERMISSIONS.EXPORT_PROYECTO_SERVICIO_SOCIAL_PDF),
        'can_export_excel': request.user.has_permission(PERMISSIONS.EXPORT_PROYECTO_SERVICIO_SOCIAL_EXCEL),
    }
    return render(request, 'servicio_list.html', context) 


@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.EXPORT_PROYECTO_SERVICIO_SOCIAL_PDF), login_url='/no-autorizado/')
def export_servicios_pdf(request):
    servicios = ServicioSocial.objects.all().order_by('-periodo_academico__fecha_inicio').prefetch_related('estudiantes_participantes')

    filter_nombre_proyecto = request.GET.get('nombre_proyecto')
    filter_tutor_encargado = request.GET.get('tutor_encargado')
    filter_estado = request.GET.get('estado')
    filter_area_accion = request.GET.get('area_accion')
    filter_periodo_id = request.GET.get('periodo')

    if filter_nombre_proyecto:
        servicios = servicios.filter(nombre_proyecto__icontains=filter_nombre_proyecto)
    if filter_tutor_encargado:
        servicios = servicios.filter(
            Q(tutor_nombres__icontains=filter_tutor_encargado) |
            Q(tutor_apellidos__icontains=filter_tutor_encargado)
        )
    if filter_estado:
        servicios = servicios.filter(estado=filter_estado)
    if filter_area_accion:
        servicios = servicios.filter(area_accion_proyecto=filter_area_accion)
    
    periodo_seleccionado = None
    if filter_periodo_id:
        servicios = servicios.filter(periodo_academico__id=filter_periodo_id)
        try:
            periodo_seleccionado = Periodo.objects.get(id=filter_periodo_id)
        except Periodo.DoesNotExist:
            pass 

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4) 
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['h1'],
        fontSize=18,
        alignment=1, # Center
        spaceAfter=14
    )
    subtitle_style = ParagraphStyle(
        name='SubtitleStyle',
        parent=styles['h2'],
        fontSize=12,
        alignment=0, # Left
        spaceAfter=8
    )

    # Nuevo estilo para elementos de lista con sangría
    indented_normal_style = ParagraphStyle(
        name='IndentedNormal',
        parent=styles['Normal'],
        leftIndent=20 # Sangría de 20 puntos
    )

    elements = []

    # --- Membrete del PDF ---
    def header_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        # Logos - Lógica comentada si no se usarán
        # path_to_logo_left = os.path.join(settings.BASE_DIR, 'servicio_social', 'static', 'img', 'escudo_izquierdo.jpg')
        # path_to_logo_right = os.path.join(settings.BASE_DIR, 'servicio_social', 'static', 'img', 'escudo_derecho.jpg')
        # try:
        #     if os.path.exists(path_to_logo_left):
        #         canvas_obj.drawImage(path_to_logo_left, inch, A4[1] - inch - 50, width=50, height=50)
        #     else:
        #         print(f"ADVERTENCIA: No se encontró el logo izquierdo para PDF en {path_to_logo_left}")
        # except Exception as e:
        #     print(f"ADVERTENCIA: Error al insertar el logo izquierdo en PDF: {e}")
        # try:
        #     if os.path.exists(path_to_logo_right):
        #         canvas_obj.drawImage(path_to_logo_right, A4[0] - inch - 50, A4[1] - inch - 50, width=50, height=50)
        #     else:
        #         print(f"ADVERTENCIA: No se encontró el logo derecho para PDF en {path_to_logo_right}")
        # except Exception as e:
        #     print(f"ADVERTENCIA: Error al insertar el logo derecho en PDF: {e}")

        # Títulos de la Universidad
        canvas_obj.setFont('Helvetica-Bold', 12)
        canvas_obj.drawCentredString(A4[0]/2.0, A4[1] - inch + 30, "UNIVERSIDAD NACIONAL EXPERIMENTAL POLITÉCNICA")
        canvas_obj.drawCentredString(A4[0]/2.0, A4[1] - inch + 15, "DE LA FUERZA ARMADA NACIONAL BOLIVARIANA")
        canvas_obj.drawCentredString(A4[0]/2.0, A4[1] - inch, "VICERRECTORADO DE ASUNTOS ACADÉMICOS")
        canvas_obj.drawCentredString(A4[0]/2.0, A4[1] - inch - 15, "UNIDAD DE SERVICIO COMUNITARIO")
        # Nuevo: Núcleo y Extensión
        canvas_obj.setFont('Helvetica', 10) # Fuente un poco más pequeña para detalles
        canvas_obj.drawCentredString(A4[0]/2.0, A4[1] - inch - 30, "Núcleo: Falcón")
        canvas_obj.drawCentredString(A4[0]/2.0, A4[1] - inch - 45, "Extensión: Punto Fijo")
        
        # Pie de página (número de página)
        canvas_obj.setFont('Helvetica', 9)
        canvas_obj.drawCentredString(A4[0]/2.0, inch/2.0, f"Página {doc_obj.page}")
        canvas_obj.restoreState()

    filtro_info = []
    if filter_nombre_proyecto: filtro_info.append(f"Proyecto: {filter_nombre_proyecto}")
    if filter_tutor_encargado: filtro_info.append(f"Tutor: {filter_tutor_encargado}")
    if filter_estado: filtro_info.append(f"Estado: {dict(ServicioSocial.ESTADO_CHOICES).get(filter_estado, filter_estado)}")
    if filter_area_accion: filtro_info.append(f"Área: {dict(ServicioSocial.AREA_ACCION_CHOICES).get(filter_area_accion, filter_area_accion)}")
    if filter_periodo_id and periodo_seleccionado: filtro_info.append(f"Período: {periodo_seleccionado.nombre}") 

    if filtro_info:
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph("Filtros Aplicados:", subtitle_style))
        for f_text in filtro_info:
            elements.append(Paragraph(f_text, styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))
        
    elements.append(Spacer(1, 0.2 * inch))
    
    proyectos_por_estado = defaultdict(int)
    for servicio in servicios:
        proyectos_por_estado[servicio.get_estado_display()] += 1
    
    proyectos_por_area = defaultdict(int)
    for servicio in servicios:
        proyectos_por_area[servicio.get_area_accion_proyecto_display()] += 1

    elements.append(Spacer(1, 0.4 * inch)) 

    # --- Tabla Principal de Datos para PDF ---
    data = [
        ["Proyecto", "Tutor", "# Estudiantes", "Estado", "Período", "Comunidad"] 
    ]
    
    for servicio in servicios:
        data.append([
            servicio.nombre_proyecto,
            f"{servicio.tutor_nombres} {servicio.tutor_apellidos}",
            str(servicio.estudiantes_participantes.count()),
            servicio.get_estado_display(),
            servicio.periodo_academico.nombre if servicio.periodo_academico else 'N/A', 
            servicio.nombre_comunidad_institucion,
        ])

    table_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')), 
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige), 
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), 
    ])

    col_widths = [3.5*cm, 3.5*cm, 1.5*cm, 2*cm, 2.5*cm, 3.5*cm] 
    table = Table(data, colWidths=col_widths)
    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer) 

    buffer.seek(0)
    return HttpResponse(buffer.getvalue(), content_type='application/pdf')

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.EXPORT_PROYECTO_SERVICIO_SOCIAL_EXCEL), login_url='/no-autorizado/')
def export_servicios_excel(request):
    servicios = ServicioSocial.objects.all().order_by('-periodo_academico__fecha_inicio').prefetch_related('estudiantes_participantes')

    filter_nombre_proyecto = request.GET.get('nombre_proyecto')
    filter_tutor_encargado = request.GET.get('tutor_encargado')
    filter_estado = request.GET.get('estado')
    filter_area_accion = request.GET.get('area_accion')
    filter_periodo_id = request.GET.get('periodo')

    if filter_nombre_proyecto:
        servicios = servicios.filter(nombre_proyecto__icontains=filter_nombre_proyecto)
    if filter_tutor_encargado:
        servicios = servicios.filter(
            Q(tutor_nombres__icontains=filter_tutor_encargado) |
            Q(tutor_apellidos__icontains=filter_tutor_encargado)
        )
    if filter_estado:
        servicios = servicios.filter(estado=filter_estado)
    if filter_area_accion:
        servicios = servicios.filter(area_accion_proyecto=filter_area_accion)
    
    periodo_seleccionado = None
    if filter_periodo_id:
        servicios = servicios.filter(periodo_academico__id=filter_periodo_id)
        try:
            periodo_seleccionado = Periodo.objects.get(id=filter_periodo_id)
        except Periodo.DoesNotExist:
            pass 

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Servicios Sociales"

    # --- Configuración de Estilos ---
    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    # Cambiar el color de la franja a un gris más suave o blanco, según tu preferencia
    header_fill = PatternFill(start_color="A0A0A0", end_color="A0A0A0", fill_type="solid") # Gris medio para los encabezados
    # O si prefieres los encabezados blancos y con borde:
    # header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
    # header_font = Font(bold=True, color="000000") # Texto negro para encabezados blancos

    sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") 
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    left_aligned_text = Alignment(horizontal="left", vertical="center")
    top_aligned_text = Alignment(horizontal="left", vertical="top", wrapText=True) 

    # --- DEFINICIÓN DE HEADERS ---
    headers = [
        "Nombre del Proyecto", "Tutor Encargado", "# Estudiantes", "Estado",
        "Período Académico", "Comunidad/Institución", "Dirección",
        "Tutor Comunitario", "C.I. Tutor Comunitario", "Tel. Tutor Comunitario",
        "Beneficiados", "Vinculación Planes", "Área Acción", "Observaciones Generales",
        "Tipo Tutor", "Unidad Administrativa", "Categoría Docente",
        "Foros", "Charlas", "Jornadas", "Talleres", "Campañas",
        "Reunión Misiones", "Ferias", "Alianzas Estratégicas"
    ]

    current_row = 1 

    # --- INFORMACIÓN DE LA UNIVERSIDAD (SIN LOGOS) ---
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "UNIVERSIDAD NACIONAL EXPERIMENTAL POLITÉCNICA DE LA FUERZA ARMADA NACIONAL BOLIVARIANA"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    ws[f'A{current_row}'].alignment = center_aligned_text
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "VICERRECTORADO DE ASUNTOS ACADÉMICOS - UNIDAD DE SERVICIO COMUNITARIO"
    ws[f'A{current_row}'].font = Font(size=12, bold=True)
    ws[f'A{current_row}'].alignment = center_aligned_text
    current_row += 1

    # Nuevo: Núcleo y Extensión para Excel
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "Núcleo: Falcón - Extensión: Punto Fijo"
    ws[f'A{current_row}'].font = Font(size=10, bold=True)
    ws[f'A{current_row}'].alignment = center_aligned_text
    current_row += 1

    current_row += 1 # Fila vacía para separación

    # Título del Reporte
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "REPORTE DETALLADO DE PROYECTOS DE SERVICIO COMUNITARIO"
    ws[f'A{current_row}'].font = Font(size=16, bold=True, color="000000")
    ws[f'A{current_row}'].alignment = center_aligned_text
    ws.row_dimensions[current_row].height = 30 
    current_row += 1

    # Fecha del Reporte y Filtros Aplicados
    current_row += 1 
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=int(len(headers)/2)) 
    ws[f'A{current_row}'] = f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    filtro_info_list = []
    if filter_nombre_proyecto: filtro_info_list.append(f"Proyecto: {filter_nombre_proyecto}")
    if filter_tutor_encargado: filtro_info_list.append(f"Tutor: {filter_tutor_encargado}")
    if filter_estado: filtro_info_list.append(f"Estado: {dict(ServicioSocial.ESTADO_CHOICES).get(filter_estado, filter_estado)}")
    if filter_area_accion: filtro_info_list.append(f"Área de Acción: {dict(ServicioSocial.AREA_ACCION_CHOICES).get(filter_area_accion, filter_area_accion)}")
    if periodo_seleccionado: filtro_info_list.append(f"Período Académico: {periodo_seleccionado.nombre}")

    if filtro_info_list:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=int(len(headers)/2))
        ws[f'A{current_row}'] = "Filtros Aplicados:"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        for f_text in filtro_info_list:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=int(len(headers)/2))
            ws[f'A{current_row}'] = f_text
            current_row += 1
    
    current_row += 2 

    # --- Sección de Estadísticas (Eliminada según tu petición) ---
    # Total de Proyectos, Estudiantes Participantes, Beneficiarios, etc.
    # Esta sección ha sido eliminada. Si la necesitas de vuelta o de otra forma, házmelo saber.
    #
    # La imagen que pasaste mostraba:
    # "Resumen Estadístico"
    # "Total de Proyectos:", total_proyectos
    # "Total de Estudiantes Participantes:", total_estudiantes_participantes
    # "Total de Beneficiarios:", total_beneficiados
    # "Proyectos por Estado:"
    # "- Pendiente:", 1
    # "Proyectos por Área de Acción:"
    # "- Tecnológico:", 1
    # Todo este bloque se ha eliminado.


    # --- Encabezados de la Tabla Principal de Datos ---
    ws.append(headers) 

    # Aplicar estilo a los encabezados
    header_row_index = current_row 
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_index, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_aligned_text
    current_row += 1 

    # --- Datos de la Tabla Principal ---
    for servicio in servicios:
        row_data = [
            servicio.nombre_proyecto,
            f"{servicio.tutor_nombres} {servicio.tutor_apellidos}",
            servicio.estudiantes_participantes.count(),
            servicio.get_estado_display(),
            servicio.periodo_academico.nombre if servicio.periodo_academico else 'N/A', 
            servicio.nombre_comunidad_institucion,
            servicio.direccion_comunidad,
            servicio.tutor_comunitario_nombre,
            servicio.tutor_comunitario_cedula,
            servicio.tutor_comunitario_telefono,
            servicio.cantidad_beneficiados,
            servicio.vinculacion_planes_programas,
            servicio.get_area_accion_proyecto_display(),
            servicio.observaciones,
            servicio.get_tutor_tipo_display(),
            servicio.tutor_unidad_administrativa,
            servicio.tutor_categoria_docente,
            "Sí" if servicio.act_foros else "No",
            "Sí" if servicio.act_charlas else "No",
            "Sí" if servicio.act_jornadas else "No",
            "Sí" if servicio.act_talleres else "No",
            "Sí" if servicio.act_campanas else "No",
            "Sí" if servicio.act_reunion_misiones else "No",
            "Sí" if servicio.act_ferias else "No",
            "Sí" if servicio.act_alianzas_estrategicas else "No",
        ]
        ws.append(row_data) 
        
        for col_num in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col_num).border = thin_border
            if headers[col_num - 1] == "Observaciones Generales": 
                ws.cell(row=ws.max_row, column=col_num).alignment = top_aligned_text


        if servicio.estudiantes_participantes.exists():
            current_row = ws.max_row + 1 
            ws.append([]) 
            current_row += 1
            
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
            ws.cell(row=current_row, column=1, value=f"Estudiantes Participantes del Proyecto: {servicio.nombre_proyecto}").font = Font(bold=True, size=11)
            ws.cell(row=current_row, column=1).alignment = left_aligned_text
            ws.cell(row=current_row, column=1).fill = sub_header_fill 
            current_row += 1

            student_headers = ["", "Nombres", "Apellidos", "Cédula", "Carrera", "Semestre", "Sección", "Turno", "Observaciones Estudiante"]
            ws.append(student_headers)
            for col_num in range(1, len(student_headers) + 1): 
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.font = Font(bold=True, size=9) 
                cell.fill = sub_header_fill
                cell.border = thin_border
                cell.alignment = center_aligned_text
            current_row = ws.max_row + 1 

            for estudiante in servicio.estudiantes_participantes.all():
                student_data = [
                    "", 
                    estudiante.nombres,
                    estudiante.apellidos,
                    estudiante.cedula_identidad,
                    estudiante.carrera.nombre if estudiante.carrera else '',
                    estudiante.semestre.nombre if estudiante.semestre else '',
                    estudiante.seccion,
                    estudiante.get_turno_display(),
                    estudiante.observaciones_estudiante,
                ]
                ws.append(student_data)
                for col_num in range(1, len(student_data) + 1): 
                    ws.cell(row=ws.max_row, column=col_num).border = thin_border
                    if student_headers[col_num - 1] == "Observaciones Estudiante":
                        ws.cell(row=ws.max_row, column=col_num).alignment = top_aligned_text

            ws.append([]) 
            current_row = ws.max_row + 1 

    # --- Autoajustar ancho de columnas ---
    # El rango de autoajuste ahora empieza después del encabezado superior,
    # que ahora ocupa las filas 1-4 (incluyendo la línea de Núcleo/Extensión y un espacio).
    # Ajusta 'initial_header_rows' si el número de filas del membrete cambia.
    initial_header_rows = 4 # Filas 1, 2, 3 (universidad/unidad) + 4 (núcleo/extensión)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx) 
        
        for row_idx in range(initial_header_rows + 1, ws.max_row + 1): # Empieza a iterar después del encabezado
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell_value = str(cell.value)
                if '\n' in cell_value:
                    lines = cell_value.split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
                else:
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
            except:
                pass
        
        adjusted_width = (max_length + 2) 
        if adjusted_width < 10: adjusted_width = 10 
        
        # Límite específico para la columna de Observaciones Generales
        # Asegúrate de que headers[col_idx-1] es válido.
        if (col_idx - 1) < len(headers) and headers[col_idx - 1] == "Observaciones Generales":
            if adjusted_width > 60:
                adjusted_width = 60 
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # La altura de la fila 1 ya no necesita ser fija para logos, pero puede serlo para el texto
    ws.row_dimensions[1].height = 20 # Altura para la primera línea de texto del encabezado

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_servicios_sociales.xlsx"' 
    wb.save(response) 
    return response


@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_PROYECTO_SERVICIO_SOCIAL), login_url='/no-autorizado/'
)
def servicio_detail(request, pk):
    servicio = get_object_or_404(ServicioSocial, pk=pk)
    
    if not request.user.has_permission(PERMISSIONS.VIEW_PROYECTO_SERVICIO_SOCIAL, obj=servicio):
        messages.error(request, "No tienes permiso para ver los detalles de este proyecto de servicio social.")
        return redirect('servicio_social:servicio_list')

    return render(request, 'servicio_detail.html', {'servicio': servicio}) 

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.ADD_PROYECTO_SERVICIO_SOCIAL), login_url='/no-autorizado/')
def servicio_create(request):
    if request.method == 'POST':
        form = ServicioSocialForm(request.POST)
        formset = EstudianteServicioSocialFormSet(request.POST, request.FILES)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    servicio = form.save(commit=False) 
                    servicio.save()
                    
                    formset.instance = servicio
                    formset.save()
                
                messages.success(request, "Proyecto de Servicio Social creado exitosamente.")
                return redirect('servicio_social:servicio_list')
            except Exception as e:
                messages.error(request, f"Error al guardar Servicio Social y estudiantes: {e}")
                print(f"Error al guardar Servicio Social y estudiantes: {e}")
                
    else: 
        form = ServicioSocialForm()
        formset = EstudianteServicioSocialFormSet()

    return render(request, 'servicio_form.html', { 
        'form': form,
        'formset': formset,
        'is_create': True
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.CHANGE_PROYECTO_SERVICIO_SOCIAL), login_url='/no-autorizado/')
def servicio_update(request, pk):
    servicio = get_object_or_404(ServicioSocial, pk=pk)
    
    if not request.user.has_permission(PERMISSIONS.CHANGE_PROYECTO_SERVICIO_SOCIAL, obj=servicio):
        messages.error(request, "No tienes permiso para editar este proyecto de servicio social.")
        return redirect('servicio_social:servicio_list')

    if request.method == 'POST':
        form = ServicioSocialForm(request.POST, instance=servicio)
        formset = EstudianteServicioSocialFormSet(request.POST, request.FILES, instance=servicio)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    servicio = form.save() 
                    formset.instance = servicio
                    formset.save()
                
                messages.success(request, "Proyecto de Servicio Social actualizado exitosamente.")
                return redirect('servicio_social:servicio_detail', pk=servicio.pk)
            except Exception as e:
                messages.error(request, f"Error al actualizar Servicio Social y estudiantes: {e}")
                print(f"Error al actualizar Servicio Social y estudiantes: {e}")
                
    else: 
        form = ServicioSocialForm(instance=servicio)
        formset = EstudianteServicioSocialFormSet(instance=servicio)
        
    return render(request, 'servicio_form.html', { 
        'form': form,
        'formset': formset,
        'servicio': servicio
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.DELETE_PROYECTO_SERVICIO_SOCIAL), login_url='/no-autorizado/')
def servicio_delete(request, pk):
    servicio = get_object_or_404(ServicioSocial, pk=pk)
    
    if not request.user.has_permission(PERMISSIONS.DELETE_PROYECTO_SERVICIO_SOCIAL, obj=servicio):
        messages.error(request, "No tienes permiso para eliminar este proyecto de servicio social.")
        return redirect('servicio_social:servicio_list')

    if request.method == 'POST':
        try:
            servicio.delete()
            messages.success(request, "Proyecto de Servicio Social eliminado exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al eliminar el proyecto de servicio social: {e}")
            print(f"Error al eliminar el proyecto de servicio social: {e}")
        return redirect('servicio_social:servicio_list')
    return render(request, 'servicio_confirm_delete.html', {'servicio': servicio}) 
