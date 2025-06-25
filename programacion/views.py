from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from django.db import IntegrityError, transaction
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from datetime import time, date
from .models import ProgramacionAcademica, Docente, Carrera, Asignatura, Periodo, Aula, HorarioAula, Seccion, HorarioSeccion, HorarioSeccion, semestre, Departamento
from .forms import ProgramacionAcademicaForm, DocenteForm, AsignaturaForm, AsignarAsignaturasForm, AulaForm, HorarioAulaForm, SeleccionarSeccionForm, SeccionForm, HorarioSeccionForm, HorarioAulaBloqueForm, HorarioAulaForm, ProgramacionAcademicaAssignmentForm, AsignaturaModalForm
from django.forms import modelformset_factory
from datetime import datetime
from django.core.exceptions import ValidationError
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from accounts.models import Usuario
from administrador.permissions import PERMISSIONS
from .utils import tiene_permiso_departal_o_carrera_util
from django.core.exceptions import PermissionDenied
from django.forms import formset_factory
# Importar openpyxl y sus componentes para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict 

# Importar ReportLab y sus componentes para PDF (se mantiene la estructura)
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from django.utils import timezone 

# Vista para la evaluación docente
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_EVALUACION_DOCENTE), login_url='/no-autorizado/')
def evaluacion_docente_list(request): # Renombrada de 'evaluacion_docente'
    query = request.GET.get('q', '')
    docentes_list = Docente.objects.all().order_by('nombre')

    if query:
        docentes_list = docentes_list.filter(
            Q(nombre__icontains=query) | 
            Q(cedula__icontains=query)
        )
    
    context = {
        'docentes': docentes_list,
        'query': query,
        'can_view_detalle_evaluacion': request.user.has_permission(PERMISSIONS.VIEW_EVALUACION_DOCENTE),
    }
    
    # La impresión de permisos se mantiene para depuración, si la necesitas.
    # print("Permisos del usuario:", request.user.get_all_permissions())
    return render(request, 'evaluacion_docente_list.html', context) # La plantilla aún se llama evaluacion_docente.html

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_EVALUACION_DOCENTE), login_url='/no-autorizado/')
def evaluacion_docente_detalle(request, docente_id): # Nueva vista de detalle
    docente = get_object_or_404(Docente, id=docente_id)
    
    filter_periodo_id = request.GET.get('periodo')
    periodos = Periodo.objects.all().order_by('-fecha_inicio')

    evaluaciones_query = ProgramacionAcademica.objects.filter(docente=docente).select_related(
        'asignatura', 'periodo', 'asignatura__carrera', 'asignatura__semestre', 'docente_evaluador'
    )

    if filter_periodo_id:
        evaluaciones_query = evaluaciones_query.filter(periodo__id=filter_periodo_id)
    
    # Ordenar las evaluaciones para la tabla
    evaluaciones = evaluaciones_query.order_by(
        'periodo__fecha_inicio',
        'asignatura__carrera__nombre',
        'asignatura__semestre__nombre',
        'asignatura__nombre'
    )

    context = {
        'docente': docente,
        'evaluaciones': evaluaciones,
        'periodos': periodos,
        'filter_periodo_id': filter_periodo_id,
        'can_manage_evaluacion': request.user.has_permission(PERMISSIONS.MANAGE_EVALUACION_DOCENTE),
        # Puedes añadir otros permisos específicos si los necesitas para la visibilidad de columnas
        'can_view_full_details': request.user.has_permission(PERMISSIONS.MANAGE_EVALUACION_DOCENTE), # O un permiso más específico si solo algunos roles ven todo
    }
    return render(request, 'evaluacion_docente_detalle.html', context)

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_EVALUACION_DOCENTE), login_url='/no-autorizado/')
def editar_evaluacion_docente(request, pk):
    evaluacion = get_object_or_404(ProgramacionAcademica, pk=pk)
    
    if request.method == 'POST':
        form = ProgramacionAcademicaForm(request.POST, instance=evaluacion)
        if form.is_valid():
            form.save()
            messages.success(request, f"Evaluación de {evaluacion.docente.nombre} para {evaluacion.asignatura.nombre} actualizada exitosamente.")
            return redirect('programacion:evaluacion_docente_detalle', docente_id=evaluacion.docente.id)
        else:
            messages.error(request, "Hubo errores al actualizar la evaluación. Por favor, revise los datos.")
            # --- LÍNEA DE DEPURACIÓN AÑADIDA ---
            print("Errores del formulario en editar_evaluacion_docente:")
            print(form.errors)
            # --- FIN LÍNEA DE DEPURACIÓN ---
            
    else:
        form = ProgramacionAcademicaForm(instance=evaluacion) 

    context = {
        'form': form,
        'evaluacion': evaluacion,
        'can_change_evaluacion': request.user.has_permission(PERMISSIONS.MANAGE_EVALUACION_DOCENTE),
    }
    return render(request, 'editar_evaluacion_docente.html', context)


@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_EVALUACION_DOCENTE), login_url='/no-autorizado/')
def eliminar_evaluacion_docente(request, pk):
    """
    Elimina una instancia específica de ProgramacionAcademica.
    """
    evaluacion = get_object_or_404(ProgramacionAcademica, pk=pk)
    docente_id = evaluacion.docente.id # Capturar el ID del docente antes de eliminar la evaluación
    
    if request.method == 'POST':
        try:
            docente_nombre = evaluacion.docente.nombre if evaluacion.docente else "N/A"
            asignatura_nombre = evaluacion.asignatura.nombre if evaluacion.asignatura else "N/A"
            evaluacion.delete()
            messages.success(request, f"La evaluación de {docente_nombre} para {asignatura_nombre} ha sido eliminada exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al eliminar la evaluación: {e}")
    
    # Redireccionar de vuelta a la página de detalle del docente
    return redirect('programacion:evaluacion_docente_detalle', docente_id=docente_id)



# Vista para listar los DOCENTES y aplicar filtros
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_DOCENTE), login_url='/no-autorizado/')
def docentes(request):
    query = request.GET.get('q', '')
    carrera_id = request.GET.get('carrera')
    dedicacion_filter = request.GET.get('dedicacion') 
    periodo_filter_id = request.GET.get('periodo') # Filtro para ProgramacionAcademica

    docentes_list = Docente.objects.all()

    if query:
        docentes_list = docentes_list.filter(
            Q(nombre__icontains=query) | 
            Q(cedula__icontains=query) |
            Q(email__icontains=query) |
            Q(dedicacion__icontains=query)
        )
    
    if carrera_id:
        docentes_list = docentes_list.filter(carreras__id=carrera_id)
    
    if dedicacion_filter:
        docentes_list = docentes_list.filter(dedicacion__iexact=dedicacion_filter) 
    
    carreras = Carrera.objects.all().order_by('nombre')
    dedicaciones_choices = sorted(list(Docente.objects.values_list('dedicacion', flat=True).distinct()))

    periodos = Periodo.objects.all().order_by('-fecha_inicio') 
    
    # Obtener las asignaciones de ProgramacionAcademica para cada docente
    docentes_with_assignments = []
    # Usamos .distinct() al final si el filtro de carrera pudiera duplicar docentes
    for docente in docentes_list.order_by('nombre'): 
        assigned_subjects_query = ProgramacionAcademica.objects.filter(docente=docente)
        
        if periodo_filter_id: # Si se está filtrando por período, aplicar a las asignaciones
            assigned_subjects_query = assigned_subjects_query.filter(periodo__id=periodo_filter_id)

        assigned_subjects = assigned_subjects_query.select_related(
            'asignatura', 'periodo', 'asignatura__carrera', 'asignatura__semestre'
        ).order_by(
            'periodo__fecha_inicio', 
            'asignatura__carrera__nombre',
            'asignatura__semestre__nombre',
            'asignatura__nombre'
        )
        docente.assigned_subjects = assigned_subjects 
        docentes_with_assignments.append(docente)

    return render(request, 'docentes.html', {
        'docentes': docentes_with_assignments, 
        'carreras': carreras,
        'dedicaciones_choices': dedicaciones_choices,
        'filter_carrera_id': carrera_id,
        'filter_dedicacion': dedicacion_filter,
        'filter_periodo_id': periodo_filter_id, 
        'query': query,
        'periodos': periodos, 
        'can_view_docente': request.user.has_permission(PERMISSIONS.VIEW_DOCENTE),
        'can_add_docente': request.user.has_permission(PERMISSIONS.MANAGE_DOCENTE), 
        'can_change_docente': request.user.has_permission(PERMISSIONS.MANAGE_DOCENTE), 
        'can_delete_docente': request.user.has_permission(PERMISSIONS.MANAGE_DOCENTE), 
        'can_assign_asignaturas': request.user.has_permission(PERMISSIONS.MANAGE_DOCENTE), 
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_DOCENTE), login_url='/no-autorizado/')
def detalle_docente(request, docente_id):
    docente = get_object_or_404(Docente, id=docente_id)
    assigned_subjects = ProgramacionAcademica.objects.filter(docente=docente).select_related(
        'asignatura', 'periodo', 'asignatura__carrera', 'asignatura__semestre'
    ).order_by(
        'periodo__fecha_inicio',
        'asignatura__carrera__nombre',
        'asignatura__semestre__nombre',
        'asignatura__nombre'
    )
    return render(request, 'detalle_docente.html', {
        'docente': docente,
        'assigned_subjects': assigned_subjects
    })

# Vista para agregar un nuevo docente
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_DOCENTE), login_url='/no-autorizado/')
def agregar_docente(request):
    if request.method == 'POST':
        form = DocenteForm(request.POST)
        if form.is_valid():
            docente = form.save()
            return redirect('programacion:docentes')
    else:
        form = DocenteForm()
    return render(request, 'agregar_docente.html', {'form': form})

# Vista para editar un docente existente
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_DOCENTE), login_url='/no-autorizado/')
def editar_docente(request, pk): 
    docente = get_object_or_404(Docente, pk=pk)
    if request.method == 'POST':
        form = DocenteForm(request.POST, instance=docente)
        if form.is_valid():
            form.save()
            return redirect('programacion:docentes')
    else:
        form = DocenteForm(instance=docente)
    return render(request, 'editar_docente.html', {'form': form, 'docente': docente})

# Vista para eliminar un docente (maneja GET para confirmación y POST para eliminación)
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_DOCENTE), login_url='/no-autorizado/')
def eliminar_docente(request, pk): 
    docente = get_object_or_404(Docente, pk=pk)
    if request.method == 'POST':
        docente.delete()
        return redirect('programacion:docentes')
    return render(request, 'docente_confirm_delete.html', {'docente': docente})


# --- VISTAS PARA ASIGNAR ASIGNATURAS A DOCENTES ---
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_DOCENTE), login_url='/no-autorizado/')
def asignar_asignaturas(request, pk):
    docente = get_object_or_404(Docente, pk=pk)
    
    # Obtener el período seleccionado para el filtro principal
    selected_period_id = request.GET.get('periodo')
    if not selected_period_id and Periodo.objects.exists():
        selected_period_id = Periodo.objects.first().id # Seleccionar el primer período si no hay uno
    
    # Formulario principal (solo para seleccionar el período)
    main_form = AsignarAsignaturasForm(request.GET or None, initial={'periodo': selected_period_id})

    # Formulario para el modal (seleccion de Carrera/Semestre/Asignatura)
    # Se pasa el objeto docente para que el queryset de Carrera se filtre correctamente
    modal_form = AsignaturaModalForm(initial={
        'carrera': request.GET.get('carrera'), # Pre-seleccionar si vienen de GET
        'semestre': request.GET.get('semestre'),
    }, docente=docente)


    # Obtener las asignaturas ya asignadas al docente para el período seleccionado
    assigned_subjects = []
    if selected_period_id and selected_period_id:
        assigned_subjects = ProgramacionAcademica.objects.filter(
            docente=docente,
            periodo__id=selected_period_id
        ).select_related('asignatura', 'periodo', 'asignatura__carrera', 'asignatura__semestre').order_by(
            'periodo__fecha_inicio',
            'asignatura__carrera__nombre',
            'asignatura__semestre__nombre',
            'asignatura__nombre'
        )

    # Si la solicitud es POST, significa que el usuario ha cambiado el período del filtro principal
    # o ha enviado el formulario del modal (manejado por vistas AJAX).
    # Este if request.method == 'POST' del flujo anterior se elimina o se simplifica,
    # ya que las asignaciones/desasignaciones individuales se harán por AJAX.
    if request.method == 'POST':
        # En este nuevo flujo, los POSTs a esta URL solo deberían ser para el filtro de período.
        # Si tienes lógica de guardado masivo aquí, deberías revisarla.
        # Para este ejemplo, asumimos que los guardados y eliminaciones se hacen vía AJAX.
        pass

    return render(request, 'asignar_asignaturas.html', {
        'docente': docente, 
        'main_form': main_form, # El formulario principal para el filtro de período
        'modal_form': modal_form, # El formulario para el modal de añadir asignatura
        'assigned_subjects': assigned_subjects, # Las asignaturas ya asignadas para la tabla
        'selected_period_id': selected_period_id, # Se usa en JS para recargar
        'periodos': Periodo.objects.all().order_by('-fecha_inicio'), # Para el selector principal
    })

# --- VISTAS API PARA CARGA DINÁMICA DE SEMESTRES Y ASIGNATURAS ---

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_DOCENTE), login_url='/no-autorizado/')
@require_POST
def api_add_assignment_to_docente(request, pk):
    docente = get_object_or_404(Docente, pk=pk) # Obtener el objeto docente

    # Pasar el objeto docente al formulario para que pueda filtrar las carreras correctamente
    form = AsignaturaModalForm(request.POST, docente=docente) 

    if form.is_valid():
        asignatura = form.cleaned_data['asignatura']
        # Asegurarse de obtener el período del POST, ya que el campo periodo está en el main_form
        periodo_id_from_post = request.POST.get('periodo_id')
        if not periodo_id_from_post:
            return JsonResponse({'success': False, 'message': 'ID de período no proporcionado.'}, status=400)
        
        try:
            periodo = Periodo.objects.get(id=periodo_id_from_post)
        except Periodo.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Período seleccionado no existe.'}, status=400)

        if not ProgramacionAcademica.objects.filter(
            docente=docente,
            asignatura=asignatura,
            periodo=periodo,
        ).exists():
            ProgramacionAcademica.objects.create(
                docente=docente,
                asignatura=asignatura,
                periodo=periodo,
            )
            return JsonResponse({'success': True, 'message': 'Asignatura asignada exitosamente.'})
        else:
            return JsonResponse({'success': False, 'message': 'Esta asignatura ya está asignada al docente para este período.'}, status=409) 
    else:
        # Aquí puedes añadir más depuración para ver los errores exactos
        print("Errores de validación en AsignaturaModalForm (POST):")
        for field, errors in form.errors.items():
            print(f"  {field}: {errors}")
        
        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({'success': False, 'message': 'Errores de validación', 'errors': errors}, status=400)

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_DOCENTE), login_url='/no-autorizado/')
@require_POST
def api_remove_assignment_from_docente(request, pk):
    # pk aquí es el ID del objeto ProgramacionAcademica a eliminar
    assignment = get_object_or_404(ProgramacionAcademica, pk=pk)
    
    try:
        assignment.delete()
        return JsonResponse({'success': True, 'message': 'Asignación eliminada exitosamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error al eliminar la asignación: {e}'}, status=500)

def api_semestres_por_carrera(request):
    carrera_id = request.GET.get('carrera_id')
    semestres_data = []
    if carrera_id:
        semestres_queryset = semestre.objects.filter(carrera__id=carrera_id).order_by('nombre')
        for s in semestres_queryset:
            semestres_data.append({'id': s.id, 'nombre': s.nombre})
    return JsonResponse(semestres_data, safe=False)



def api_asignaturas_por_carrera_semestre(request):
    carrera_id_str = request.GET.get('carrera_id')
    semestre_id_str = request.GET.get('semestre_id')

    print(f"\n--- DEBUG: api_asignaturas_por_carrera_semestre ---")
    print(f"Recibido carrera_id (str): '{carrera_id_str}'")
    print(f"Recibido semestre_id (str): '{semestre_id_str}'")

    asignaturas_data = []
    
    # Conversión explícita a int con manejo de errores
    _carrera_id_int = None
    _semestre_id_int = None

    if carrera_id_str:
        try:
            _carrera_id_int = int(carrera_id_str)
        except (ValueError, TypeError) as e:
            print(f"ERROR: carrera_id '{carrera_id_str}' no es un entero válido. Excepción: {e}")
            # Considera devolver un error 400 si esto es crítico para el frontend
            # return JsonResponse({'error': 'ID de carrera no válido'}, status=400)
            return JsonResponse([], safe=False) # Devuelve vacío para no romper el JS

    if semestre_id_str:
        try:
            _semestre_id_int = int(semestre_id_str)
        except (ValueError, TypeError) as e:
            print(f"ERROR: semestre_id '{semestre_id_str}' no es un entero válido. Excepción: {e}")
            # return JsonResponse({'error': 'ID de semestre no válido'}, status=400)
            return JsonResponse([], safe=False) # Devuelve vacío para no romper el JS


    print(f"Carrera ID (int): {_carrera_id_int}")
    print(f"Semestre ID (int): {_semestre_id_int}")

    # Solo si ambos IDs son enteros válidos, intentar la consulta
    if _carrera_id_int is not None and _semestre_id_int is not None:
        try:
            asignaturas_queryset = Asignatura.objects.filter(
                carrera__id=_carrera_id_int,
                semestre__id=_semestre_id_int
            ).order_by('nombre')

            print(f"Consulta a la base de datos: Asignatura.objects.filter(carrera__id={_carrera_id_int}, semestre__id={_semestre_id_int})")
            print(f"Número de asignaturas encontradas: {asignaturas_queryset.count()}")

            for a in asignaturas_queryset:
                asignaturas_data.append({
                    'id': a.id,
                    'nombre': a.nombre,
                    'codigo': a.codigo,
                })
        except Exception as e:
            print(f"ERROR: Excepción en la consulta de Asignaturas: {e}")
            # return JsonResponse({'error': f'Error en la base de datos: {e}'}, status=500)
            return JsonResponse([], safe=False) # Devuelve vacío para no romper el JS
    else:
        print("ADVERTENCIA: Faltan carrera_id o semestre_id válidos para realizar la consulta.")
        # Devuelve asignaturas_data[] que ya está vacío por defecto

    print(f"Datos de asignaturas enviados en la respuesta: {asignaturas_data}")
    print(f"---------------------------------------------\n")

    return JsonResponse(asignaturas_data, safe=False)# Vistas de ProgramacionAcademica existentes (se mantienen sin cambios importantes aquí)
# def evaluacion_docente(request): ...

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_ASIGNATURA), login_url='/no-autorizado/')
def asignaturas(request):
    query = request.GET.get('q', '')
    carrera_id = request.GET.get('carrera')
    semestre_id = request.GET.get('semestre') # Cambiamos 'semestre' a 'semestre_id' para consistencia

    asignaturas_qs = Asignatura.objects.select_related('carrera', 'semestre').all() # Optimización

    if query:
        asignaturas_qs = asignaturas_qs.filter(
            Q(nombre__icontains=query) |
            Q(codigo__icontains=query)
        )
    
    if carrera_id:
        asignaturas_qs = asignaturas_qs.filter(carrera__id=carrera_id)
    
    if semestre_id:
        asignaturas_qs = asignaturas_qs.filter(semestre__id=semestre_id) # Filtrar por ID de semestre
    
    carreras = Carrera.objects.all().order_by('nombre')
    
    # Obtener los semestres disponibles para el filtro (podrían ser todos o solo los de la carrera seleccionada)
    semestres_para_filtro = semestre.objects.all().order_by('nombre')
    if carrera_id:
        semestres_para_filtro = semestres_para_filtro.filter(carrera__id=carrera_id)


    return render(request, 'asignaturas.html', {
        'asignaturas': asignaturas_qs,
        'carreras': carreras,
        'semestres': semestres_para_filtro, # Pasar los semestres filtrados (o todos)
        'query': query,
        'filter_carrera_id': carrera_id, # Usar 'filter_carrera_id' para el template
        'filter_semestre_id': semestre_id, # Usar 'filter_semestre_id' para el template
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_EVALUACION_DOCENTE), login_url='/no-autorizado/')
def programacion_lista(request):
    programaciones = ProgramacionAcademica.objects.all()
    return render(request, 'programacion_lista.html', {'programaciones': programaciones})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_ASIGNATURA), login_url='/no-autorizado/')
def agregar_asignatura(request):
    if request.method == 'POST':
        form = AsignaturaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('programacion:asignaturas')
    else:
        # Para GET, inicializa el formulario. Pasa todas las carreras.
        # El semestre se inicializará dinámicamente con JS.
        form = AsignaturaForm()
    
    carreras = Carrera.objects.all().order_by('nombre')
    # Se pasa 'semestres' para el caso inicial o para la recarga, pero JS hará la magia.
    # Inicialmente, el queryset de semestre en el form puede estar vacío o con todos.
    semestres_qs = semestre.objects.all().order_by('nombre') 

    return render(request, 'agregar_asignatura.html', {
        'form': form,
        'carreras': carreras,
        'semestres_all': semestres_qs, # Pasa todos los semestres para que el JS pueda filtrar
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_ASIGNATURA), login_url='/no-autorizado/')
def editar_asignatura(request, asignatura_id):
    asignatura = get_object_or_404(Asignatura, id=asignatura_id)
    if request.method == 'POST':
        form = AsignaturaForm(request.POST, instance=asignatura)
        if form.is_valid():
            form.save()
            return redirect('programacion:asignaturas')
    else:
        form = AsignaturaForm(instance=asignatura)
    
    carreras = Carrera.objects.all().order_by('nombre')
    
    # Para editar, necesitas los semestres de la carrera actual de la asignatura
    semestres_qs = semestre.objects.all().order_by('nombre')
    if asignatura.carrera:
        semestres_qs = semestre.objects.filter(carrera=asignatura.carrera).order_by('nombre')

    return render(request, 'editar_asignatura.html', {
        'form': form, 
        'asignatura': asignatura,
        'carreras': carreras,
        'semestres_all': semestres_qs, # Pasa los semestres filtrados por la carrera de la asignatura
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_ASIGNATURA), login_url='/no-autorizado/')
def eliminar_asignatura(request, asignatura_id):
    asignatura = get_object_or_404(Asignatura, id=asignatura_id)
    if request.method == 'POST':
        asignatura.delete()
        return redirect('programacion:asignaturas')
    return render(request, 'eliminar_asignatura.html', {'asignatura': asignatura})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_ASIGNATURA), login_url='/no-autorizado/')
def detalle_asignatura(request, asignatura_id):
    asignatura = get_object_or_404(Asignatura, id=asignatura_id)
    return render(request, 'detalle_asignatura.html', {'asignatura': asignatura})

# --- Aulas ---

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_AULA), login_url='/no-autorizado/')
def aula_list(request):
    aulas = Aula.objects.all()
    return render(request, 'aula_list.html', {'aulas': aulas})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_AULA), login_url='/no-autorizado/')
def aula_create(request):
    if request.method == 'POST':
        form = AulaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('programacion:aula_list')
    else:
        form = AulaForm()
    return render(request, 'aula_form.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_AULA), login_url='/no-autorizado/')
def aula_edit(request, pk):
    aula = get_object_or_404(Aula, pk=pk)
    if request.method == 'POST':
        form = AulaForm(request.POST, instance=aula)
        if form.is_valid():
            form.save()
            return redirect('programacion:aula_list')
    else:
        form = AulaForm(instance=aula)
    return render(request, 'aula_form.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_AULA), login_url='/no-autorizado/')
def aula_delete(request, pk):
    aula = get_object_or_404(Aula, pk=pk)
    if request.method == 'POST':
        aula.delete()
        return redirect('programacion:aula_list')
    return render(request, 'aula_confirm_delete.html', {'aula': aula})

# --- Horarios ---
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_HORARIO_AULA), login_url='/no-autorizado/')
def horario_list(request):
    horarios = HorarioAula.objects.select_related('aula', 'asignatura', 'carrera').all()
    return render(request, 'horario_list.html', {'horarios': horarios})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_AULA), login_url='/no-autorizado/')
def horario_create(request):
    aulas = Aula.objects.all()
    aula_id = request.GET.get('aula')
    aula = Aula.objects.filter(id=aula_id).first() if aula_id else None

    HorarioFormSet = modelformset_factory(HorarioAula, form=HorarioAulaBloqueForm, extra=3, can_delete=True)
    if request.method == 'POST':
        formset = HorarioFormSet(request.POST, queryset=HorarioAula.objects.none())
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                    horario = form.save(commit=False)
                    horario.aula = aula  # asigna el aula seleccionada
                    horario.save()
            return redirect('programacion:horario_list')
    else:
        formset = HorarioFormSet(queryset=HorarioAula.objects.none())

    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    horas = [
        ("07:00", "07:45"),
        ("07:45", "08:30"),
        ("08:30", "09:15"),
        ("09:15", "10:00"),
        ("10:00", "10:45"),
        ("10:45", "11:30"),
        ("11:30", "12:15"),
        ("12:15", "13:00"),
        ("13:00", "13:45"),
        ("13:45", "14:30"),
        ("14:30", "15:15"),
        ("15:15", "16:00"),
        ("16:00", "16:45"),
        ("16:45", "17:30"),
        ("17:30", "18:15"),
        ("18:15", "19:00"),
        ("19:00", "19:45"),
        ("19:45", "20:30"),
        ("20:30", "21:15"),
        ("21:15", "22:00"),
    ]
    grilla = {}
    if aula_id:
        horarios = HorarioAula.objects.filter(aula=aula).select_related('asignatura', 'docente')
        for h in horarios:
            for hora_inicio_str, hora_fin_str in horas:
                hora_inicio_dt = datetime.strptime(hora_inicio_str, "%H:%M").time()
                hora_fin_dt = datetime.strptime(hora_fin_str, "%H:%M").time()
                if (
                    h.dia in dias and
                    h.hora_inicio < hora_fin_dt and
                    h.hora_fin > hora_inicio_dt
                ):
                    grilla[(h.aula_id, h.dia, hora_inicio_str)] = h

    return render(request, 'horario_form.html', {
        'formset': formset,
        'aula': aula,
        'aulas': aulas,  # <-- agrega esto
        'dias': dias,
        'horas': horas,
        'grilla': grilla,
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_AULA), login_url='/no-autorizado/')
def horario_edit(request, pk):
    horario = get_object_or_404(HorarioAula, pk=pk)
    if request.method == 'POST':
        form = HorarioAulaForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            return redirect('programacion:horario_list')
    else:
        form = HorarioAulaForm(instance=horario)
    return render(request, 'horario_form.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_AULA), login_url='/no-autorizado/')
@require_POST # Asegura que solo acepte peticiones POST
def horario_delete(request, pk):
    horario = get_object_or_404(HorarioAula, pk=pk)
    
    # Comprobar si la petición es AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            horario.delete()
            return JsonResponse({'success': True, 'message': 'Bloque de horario eliminado exitosamente.'})
        except Exception as e:
            print(f"Error al eliminar el bloque de horario por AJAX: {e}")
            return JsonResponse({'success': False, 'message': f'Error al eliminar el bloque: {e}'}, status=500)
    
    # Si no es una petición AJAX (ej. formulario normal de confirmación de borrado)
    if request.method == 'POST':
        horario.delete()
        return redirect('programacion:horario_list')
    
    # Para peticiones GET no AJAX, mostrar página de confirmación
    return render(request, 'horario_confirm_delete.html', {'horario': horario})

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_HORARIO_AULA), login_url='/no-autorizado/')
def grilla_aulario(request):
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    horas = [
        ("07:00", "07:45"), ("07:45", "08:30"), ("08:30", "09:15"), ("09:15", "10:00"),
        ("10:00", "10:45"), ("10:45", "11:30"), ("11:30", "12:15"), ("12:15", "13:00"),
        ("13:00", "13:45"), ("13:45", "14:30"), ("14:30", "15:15"), ("15:15", "16:00"),
        ("16:00", "16:45"), ("16:45", "17:30"), ("17:30", "18:15"), ("18:15", "19:00"),
        ("19:00", "19:45"), ("19:45", "20:30"), ("20:30", "21:15"), ("21:15", "22:00"),
    ]

    # aulas = Aula.objects.all().order_by('nombre') # ¡Eliminamos esta línea o la modificamos más abajo!
    
    # --- Lógica de Filtros ---
    docente_filter_id = request.GET.get('docente')
    seccion_filter_id = request.GET.get('seccion')
    semestre_filter_id = request.GET.get('semestre')
    carrera_filter_id = request.GET.get('carrera')

    # Consulta base para todos los bloques de horarioAula
    bloques_horario_queryset = HorarioAula.objects.select_related(
        'asignatura', 'aula', 'docente', 'horario_seccion__seccion',
        'horario_seccion__seccion__carrera', 'horario_seccion__seccion__semestre'
    ).all()

    # Aplicar filtros si están presentes
    if docente_filter_id:
        bloques_horario_queryset = bloques_horario_queryset.filter(docente__id=docente_filter_id)
    if seccion_filter_id:
        bloques_horario_queryset = bloques_horario_queryset.filter(horario_seccion__seccion__id=seccion_filter_id)
    if semestre_filter_id:
        bloques_horario_queryset = bloques_horario_queryset.filter(horario_seccion__seccion__semestre__id=semestre_filter_id)
    if carrera_filter_id:
        bloques_horario_queryset = bloques_horario_queryset.filter(horario_seccion__seccion__carrera__id=carrera_filter_id)

    # Ordenar los bloques horarios para asegurar consistencia en la grilla
    bloques_horario_queryset = bloques_horario_queryset.order_by(
        'aula__nombre', 'dia', 'hora_inicio'
    )

    grilla = {}
    aula_ids_con_horarios_filtrados = set() # Conjunto para almacenar IDs de aulas con horarios filtrados

    for bloque in bloques_horario_queryset: # Usar el queryset filtrado
        # Formato de clave: "aula_id-Dia-HH:MM"
        clave = f"{bloque.aula.id}-{bloque.dia}-{bloque.hora_inicio.strftime('%H:%M')}"
        grilla[clave] = bloque
        aula_ids_con_horarios_filtrados.add(bloque.aula.id) # Añadir el ID del aula

    # Ahora, filtra la lista de aulas para mostrar solo aquellas que tienen bloques horarios que coinciden con los filtros
    if aula_ids_con_horarios_filtrados:
        aulas = Aula.objects.filter(id__in=aula_ids_con_horarios_filtrados).order_by('nombre')
    else:
        # Si no hay bloques horarios que coincidan con los filtros, no se muestran aulas.
        # Esto significa que el filtro es tan restrictivo que ninguna aula tiene horarios con esos criterios.
        aulas = Aula.objects.none() 


    # Obtener todas las opciones para los filtros (se muestran siempre en los selects)
    docentes = Docente.objects.all().order_by('nombre')
    secciones = Seccion.objects.all().order_by('codigo')
    semestres = semestre.objects.all().order_by('nombre')
    carreras = Carrera.objects.all().order_by('nombre')

    return render(request, 'grilla_aulario.html', {
        'aulas': aulas, # Ahora 'aulas' contiene solo las aulas con horarios filtrados
        'dias': dias,
        'horas': horas,
        'grilla': grilla,
        'docentes': docentes,
        'secciones': secciones,
        'semestres': semestres,
        'carreras': carreras,
        'selected_docente': docente_filter_id,
        'selected_seccion': seccion_filter_id,
        'selected_semestre': semestre_filter_id,
        'selected_carrera': carrera_filter_id,
    })
    
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_SECCION), login_url='/no-autorizado/')
def seccion_list(request):
    query = request.GET.get('q', '')
    carrera_id = request.GET.get('carrera')
    semestre_id = request.GET.get('semestre')
    periodo_id = request.GET.get('periodo') # Nuevo filtro por período

    secciones_qs = Seccion.objects.select_related('carrera', 'semestre', 'periodo').all() # Optimización

    if query:
        secciones_qs = secciones_qs.filter(
            Q(codigo__icontains=query) |
            Q(nombre__icontains=query)
        )
    
    if carrera_id:
        secciones_qs = secciones_qs.filter(carrera__id=carrera_id)
    
    if semestre_id:
        secciones_qs = secciones_qs.filter(semestre__id=semestre_id)
    
    if periodo_id: # Aplicar filtro por período
        secciones_qs = secciones_qs.filter(periodo__id=periodo_id)
    
    carreras = Carrera.objects.all().order_by('nombre')
    periodos = Periodo.objects.all().order_by('-fecha_inicio')

    # Semestres para el filtro (filtrados por carrera si hay una seleccionada)
    semestres_para_filtro = semestre.objects.all().order_by('nombre')
    if carrera_id:
        semestres_para_filtro = semestres_para_filtro.filter(carrera__id=carrera_id)

    context = {
        'secciones': secciones_qs,
        'carreras': carreras,
        'semestres': semestres_para_filtro,
        'periodos': periodos,
        'query': query,
        'filter_carrera_id': carrera_id,
        'filter_semestre_id': semestre_id,
        'filter_periodo_id': periodo_id,
        # Variables de permiso a pasar al template:
        'can_create_seccion': request.user.has_permission(PERMISSIONS.MANAGE_SECCION),
        'can_edit_seccion': request.user.has_permission(PERMISSIONS.MANAGE_SECCION),
        'can_delete_seccion': request.user.has_permission(PERMISSIONS.MANAGE_SECCION),
        'can_program_horario': request.user.has_permission(PERMISSIONS.MANAGE_HORARIO_SECCION), # Asumo este permiso para programar
    }
    return render(request, 'seccion_list.html', context) # Ajusta la ruta si es necesario

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_SECCION), login_url='/no-autorizado/')
def seccion_create(request):
    if request.method == 'POST':
        form = SeccionForm(request.POST)
        if form.is_valid():
            seccion = form.save(commit=False)
            # Aplicar granularidad al crear: El usuario solo puede crear secciones en su ámbito
            if not request.user.has_permission(PERMISSIONS.MANAGE_SECCION, obj=seccion):
                messages.error(request, "No tienes permiso para crear esta sección en la carrera/departamento especificado.")
                return redirect('programacion:seccion_list') # Redirige o muestra error

            seccion.save()
            messages.success(request, "Sección creada exitosamente.")
            return redirect('programacion:seccion_list')
    else:
        form = SeccionForm()
    
    carreras = Carrera.objects.all().order_by('nombre')
    periodos = Periodo.objects.all().order_by('-fecha_inicio')
    semestres_all = semestre.objects.all().order_by('nombre') # Corregido a Semestre

    context = {
        'form': form,
        'carreras': carreras,
        'periodos': periodos,
        'semestres_all': semestres_all,
    }
    return render(request, 'seccion_form.html', context) # Ajusta la ruta

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_SECCION), login_url='/no-autorizado/')
def seccion_edit(request, pk):
    seccion = get_object_or_404(Seccion, pk=pk)
    # Aplicar granularidad al editar: El usuario solo puede editar secciones en su ámbito
    if not request.user.has_permission(PERMISSIONS.MANAGE_SECCION, obj=seccion):
        messages.error(request, "No tienes permiso para editar esta sección.")
        return redirect('programacion:seccion_list') # Redirige o muestra error

    if request.method == 'POST':
        form = SeccionForm(request.POST, instance=seccion)
        if form.is_valid():
            seccion_updated = form.save(commit=False)
            # También aplica la granularidad si el departamento/carrera de la sección ha cambiado
            if not request.user.has_permission(PERMISSIONS.MANAGE_SECCION, obj=seccion_updated):
                messages.error(request, "No tienes permiso para mover esta sección a la nueva carrera/departamento.")
                return redirect('programacion:seccion_list') # Redirige o muestra error

            seccion_updated.save()
            messages.success(request, "Sección actualizada exitosamente.")
            return redirect('programacion:seccion_list')
    else:
        form = SeccionForm(instance=seccion)
    
    carreras = Carrera.objects.all().order_by('nombre')
    periodos = Periodo.objects.all().order_by('-fecha_inicio')
    
    semestres_all = semestre.objects.all().order_by('nombre') # Corregido a Semestre
    if seccion.carrera:
        semestres_all = semestre.objects.filter(carrera=seccion.carrera).order_by('nombre') # Corregido a Semestre

    context = {
        'form': form, 
        'seccion': seccion,
        'carreras': carreras,
        'periodos': periodos,
        'semestres_all': semestres_all,
    }
    return render(request, 'seccion_form.html', context) # Ajusta la ruta

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_SECCION), login_url='/no-autorizado/')
def seccion_delete(request, pk):
    seccion = get_object_or_404(Seccion, pk=pk)
    # Aplicar granularidad al eliminar
    if not request.user.has_permission(PERMISSIONS.MANAGE_SECCION, obj=seccion):
        messages.error(request, "No tienes permiso para eliminar esta sección.")
        return redirect('programacion:seccion_list') # Redirige o muestra error

    if request.method == 'POST':
        try:
            seccion.delete()
            messages.success(request, "Sección eliminada exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al eliminar la sección: {e}")
        return redirect('programacion:seccion_list')
    return render(request, 'seccion_confirm_delete.html', {'seccion': seccion}) # Ajusta la ruta
    
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_SECCION), login_url='/no-autorizado/')
def seleccionar_seccion(request):
    query = request.GET.get('q', '')
    carrera_id = request.GET.get('carrera')
    semestre_id = request.GET.get('semestre')
    periodo_id = request.GET.get('periodo')

    secciones_qs = Seccion.objects.select_related('carrera', 'semestre', 'periodo').all()

    if query:
        secciones_qs = secciones_qs.filter(
            Q(codigo__icontains=query) |
            Q(nombre__icontains=query)
        )
    
    if carrera_id:
        secciones_qs = secciones_qs.filter(carrera__id=carrera_id)
    
    if semestre_id:
        secciones_qs = secciones_qs.filter(semestre__id=semestre_id)
    
    if periodo_id:
        secciones_qs = secciones_qs.filter(periodo__id=periodo_id)
    
    secciones_qs = secciones_qs.order_by('carrera__nombre', 'semestre__nombre', 'codigo')

    form_initial_data = {
        'q': query,
        'carrera': carrera_id,
        'semestre': semestre_id,
        'periodo': periodo_id,
    }

    form = SeleccionarSeccionForm(
        initial=form_initial_data,
        secciones_queryset=secciones_qs 
    )

    carreras = Carrera.objects.all().order_by('nombre')
    periodos = Periodo.objects.all().order_by('-fecha_inicio')
    
    semestres_para_filtro = semestre.objects.all().order_by('nombre')
    if carrera_id:
        semestres_para_filtro = semestres_para_filtro.filter(carrera__id=carrera_id)

    return render(request, 'seleccionar_seccion.html', {
        'form': form,
        'carreras': carreras, 
        'semestres': semestres_para_filtro, 
        'periodos': periodos, 
        'filter_q': query,
        'filter_carrera_id': carrera_id,
        'filter_semestre_id': semestre_id,
        'filter_periodo_id': periodo_id,
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_HORARIO_SECCION), login_url='/no-autorizado/')
def aulario_dashboard(request):
    # Obtener todos los horarios de sección activos
    # Estos representan las "grillas activas" de las que hablas
    horarios_activos = HorarioSeccion.objects.filter(activo=True).select_related('seccion__carrera', 'seccion__semestre', 'periodo').order_by('seccion__carrera__nombre', 'seccion__semestre__nombre', 'seccion__codigo')
    
    # También puedes pasar todas las secciones para la opción de crear nuevo aulario si es necesario
    secciones = Seccion.objects.all().select_related('carrera', 'semestre').order_by('carrera__nombre', 'semestre__nombre', 'codigo')

    return render(request, 'aulario_dashboard.html', {
        'horarios_activos': horarios_activos,
        'secciones': secciones, # Secciones disponibles para crear un nuevo horario
    })

@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.VIEW_HORARIO_SECCION), login_url='/no-autorizado/')
def programar_horario(request, seccion_id):
    seccion = get_object_or_404(Seccion, id=seccion_id)
    horarios_seccion = seccion.horarios.order_by('fecha_inicio')
    horario_activo = horarios_seccion.filter(activo=True).first()
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    horas = [
        ("07:00", "07:45"),
        ("07:45", "08:30"),
        ("08:30", "09:15"),
        ("09:15", "10:00"),
        ("10:00", "10:45"),
        ("10:45", "11:30"),
        ("11:30", "12:15"),
        ("12:15", "13:00"),
        ("13:00", "13:45"),
        ("13:45", "14:30"),
        ("14:30", "15:15"),
        ("15:15", "16:00"),
        ("16:00", "16:45"),
        ("16:45", "17:30"),
        ("17:30", "18:15"),
        ("18:15", "19:00"),
        ("19:00", "19:45"),
        ("19:45", "20:30"),
        ("20:30", "21:15"),
        ("21:15", "22:00"),
    ]
    
    # Construir la grilla: {(dia, hora_inicio): horario}
    grilla = {}
    if horario_activo:
        horarios_aula_activos = HorarioAula.objects.filter(horario_seccion=horario_activo)
        for h in horarios_aula_activos:
            clave = f"{h.dia}-{h.hora_inicio.strftime('%H:%M')}"
            grilla[clave] = h
        
    # Filtrar asignaturas por carrera y semestre de la sección
    asignaturas = Asignatura.objects.filter(
        carrera=seccion.carrera,
        semestre=seccion.semestre 
    ).order_by('nombre')

    asignaturas_info = []
    if horario_activo:
        for asignatura in asignaturas:
            sesiones_programadas = HorarioAula.objects.filter(
                horario_seccion=horario_activo,
                asignatura=asignatura
            ).count()
            sesiones_planificadas = asignatura.horas_teoricas + asignatura.horas_practicas + asignatura.horas_laboratorio
            asignaturas_info.append({
                'asignatura': asignatura,
                'sesiones_programadas': sesiones_programadas,
                'sesiones_planificadas': sesiones_planificadas,
            })

    return render(request, 'programar_horario.html', {
        'seccion': seccion,
        'horarios_seccion': horarios_seccion,
        'horario_activo': horario_activo,
        'dias': dias,
        'horas': horas,
        'grilla': grilla,
        'asignaturas_info': asignaturas_info, # Renombrado para consistencia con el template actual
        'aulas': Aula.objects.all(),
        'docentes': Docente.objects.all(),
        'horario_seccion_form': HorarioSeccionForm(), # Renombrado el form para ser más explícito
        'asignaturas_seccion': asignaturas, # Pasar las asignaturas filtradas de la sección para el modal
    })
    
    return render(request, 'programar_horario.html', {
        'seccion': seccion,
        'horarios_seccion': horarios_seccion,
        'horario_activo': horario_activo,
        'dias': dias,
        'horas': horas,
        'grilla': grilla,
        'asignaturas_info': asignaturas_info, # Renombrado para consistencia, usaremos este en el template
        'asignaturas_seccion': asignaturas_seccion, # Para el dropdown del modal
        'aulas': Aula.objects.all().order_by('nombre'), 
        'docentes': Docente.objects.all().order_by('nombre'), 
        'horario_seccion_form': HorarioSeccionForm(), 
    })
    

@require_POST
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_AULA), login_url='/no-autorizado/')
def guardar_bloque_horario(request, seccion_id):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            dia = request.POST.get('dia')
            hora_inicio_str = request.POST.get('hora_inicio') 
            asignatura_id = request.POST.get('asignatura')
            aula_id = request.POST.get('aula')
            docente_id = request.POST.get('docente')

            # Validación: asegúrate de que todos los IDs estén presentes y sean números
            if not all([asignatura_id, aula_id, docente_id]):
                return JsonResponse({'success': False, 'message': 'Todos los campos son obligatorios.'}, status=400)
            if not (asignatura_id.isdigit() and aula_id.isdigit() and docente_id.isdigit()):
                return JsonResponse({'success': False, 'message': 'ID inválido en los campos.'}, status=400)

            from .models import Seccion, HorarioSeccion, Asignatura, Aula, Docente, HorarioAula

            seccion = get_object_or_404(Seccion, id=seccion_id)
            horario_activo = HorarioSeccion.objects.filter(seccion=seccion, activo=True).first()
            
            if not horario_activo:
                return JsonResponse({'success': False, 'message': 'No hay un horario activo para esta sección.'}, status=400)

            asignatura = get_object_or_404(Asignatura, id=asignatura_id)
            aula = get_object_or_404(Aula, id=aula_id)
            docente = get_object_or_404(Docente, id=docente_id)

            hora_inicio = time.fromisoformat(hora_inicio_str)

            # Determinar hora_fin según tu lógica de bloques
            horas_predefinidas = [
                ("07:00", "07:45"), ("07:45", "08:30"), ("08:30", "09:15"), ("09:15", "10:00"),
                ("10:00", "10:45"), ("10:45", "11:30"), ("11:30", "12:15"), ("12:15", "13:00"),
                ("13:00", "13:45"), ("13:45", "14:30"), ("14:30", "15:15"), ("15:15", "16:00"),
                ("16:00", "16:45"), ("16:45", "17:30"), ("17:30", "18:15"), ("18:15", "19:00"),
                ("19:00", "19:45"), ("19:45", "20:30"), ("20:30", "21:15"), ("21:15", "22:00"),
            ]
            hora_fin = None
            for h_ini, h_fin in horas_predefinidas:
                if h_ini == hora_inicio_str: 
                    hora_fin = time.fromisoformat(h_fin) 
                    break
            
            if hora_fin is None:
                return JsonResponse({'success': False, 'message': 'No se pudo determinar la hora de fin para la hora de inicio proporcionada.'}, status=400)

            # Validación de solapamiento de aula
            conflicto = HorarioAula.objects.filter(
                aula=aula,
                dia=dia,
                hora_inicio__lt=hora_fin,
                hora_fin__gt=hora_inicio,
                horario_seccion__activo=True
            ).exclude(horario_seccion=horario_activo).exists()

            if conflicto:
                return JsonResponse({'success': False, 'message': '¡El aula ya está ocupada en ese horario!'}, status=400)

            semestre_asignatura = asignatura.semestre

            bloque, created = HorarioAula.objects.update_or_create(
                horario_seccion=horario_activo, 
                dia=dia,                       
                hora_inicio=hora_inicio,       
                defaults={
                    'asignatura': asignatura,
                    'aula': aula,
                    'docente': docente,
                    'hora_fin': hora_fin,
                    'semestre': semestre_asignatura, 
                    'carrera': seccion.carrera,     
                    'seccion': seccion.codigo, 
                }
            )
            return JsonResponse({'success': True})

        except ValidationError as e:
            return JsonResponse({'success': False, 'message': f'Error de validación: {e.message}'}, status=400)
        except Exception as e:
            print(f"Error al guardar el bloque horario: {e}")
            return JsonResponse({'success': False, 'message': f'Error interno del servidor: {e}'}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Solicitud inválida.'}, status=400)


@require_POST
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_SECCION), login_url='/no-autorizado/')
def crear_horario_seccion(request, seccion_id):
    seccion = get_object_or_404(Seccion, id=seccion_id)
    form = HorarioSeccionForm(request.POST)
    if form.is_valid():
        HorarioSeccion.objects.filter(seccion=seccion).update(activo=False)
        horario = form.save(commit=False)
        horario.seccion = seccion
        horario.activo = True  
        horario.save()
        return JsonResponse({'success': True, 'horario_id': horario.id, 'nombre': horario.__str__(), 'fecha_inicio': horario.fecha_inicio.strftime('%d/%m/%Y'), 'fecha_fin': horario.fecha_fin.strftime('%d/%m/%Y')})
    else:
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)

@require_POST
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_SECCION), login_url='/no-autorizado/')
def editar_horario_seccion(request, horario_id):
    horario = get_object_or_404(HorarioSeccion, id=horario_id)
    if request.method == 'POST':
        form = HorarioSeccionForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'nombre': horario.__str__(), 'fecha_inicio': horario.fecha_inicio.strftime('%d/%m/%Y'), 'fecha_fin': horario.fecha_fin.strftime('%d/%m/%Y')})
        else:
            html = render_to_string('horario_seccion_edit_modal_content.html', {'form': form, 'horario': horario}, request)
            return JsonResponse({'success': False, 'form_html': html, 'errors': form.errors}, status=400)
    else:
        form = HorarioSeccionForm(instance=horario)
        html = render_to_string('horario_seccion_edit_modal_content.html', {'form': form, 'horario': horario}, request)
        return JsonResponse({'form_html': html})

@require_POST
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_SECCION), login_url='/no-autorizado/')
def editar_horario_seccion(request, horario_id):
    horario = get_object_or_404(HorarioSeccion, id=horario_id)
    if request.method == 'POST':
        form = HorarioSeccionForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'nombre': horario.__str__(), 'fecha_inicio': horario.fecha_inicio.strftime('%d/%m/%Y'), 'fecha_fin': horario.fecha_fin.strftime('%d/%m/%Y')})
        else:
            html = render_to_string('horario_seccion_edit_modal_content.html', {'form': form, 'horario': horario}, request)
            return JsonResponse({'success': False, 'form_html': html, 'errors': form.errors}, status=400)
    else:
        form = HorarioSeccionForm(instance=horario)
        html = render_to_string('horario_seccion_edit_modal_content.html', {'form': form, 'horario': horario}, request)
        return JsonResponse({'form_html': html})

@require_POST
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_SECCION), login_url='/no-autorizado/')
def activar_horario_seccion(request, horario_id):
    horario_a_activar = get_object_or_404(HorarioSeccion, id=horario_id)
    seccion_id = horario_a_activar.seccion.id
    HorarioSeccion.objects.filter(seccion=horario_a_activar.seccion).update(activo=False)
    horario_a_activar.activo = True
    horario_a_activar.save()
    return JsonResponse({'success': True, 'seccion_id': seccion_id}) 

@require_POST
@login_required
@user_passes_test(lambda u: u.has_permission(PERMISSIONS.MANAGE_HORARIO_SECCION), login_url='/no-autorizado/')
def eliminar_horario_seccion(request, horario_id):
    horario = get_object_or_404(HorarioSeccion, id=horario_id)
    seccion_id = horario.seccion.id 
    horario.delete()
    return JsonResponse({'success': True, 'seccion_id': seccion_id}) 


def ajax_semestres(request):
    carrera_id = request.GET.get('carrera_id')
    semestres = []
    if carrera_id:
        semestres = list(semestre.objects.filter(carrera_id=carrera_id).values('id', 'nombre'))
    return JsonResponse({'semestres': semestres})

def export_evaluacion_excel(request):
    evaluaciones_query = ProgramacionAcademica.objects.all().select_related(
        'docente', 'asignatura', 'periodo', 'docente_evaluador'
    ).order_by(
        '-periodo__fecha_inicio', 'docente__nombre', 'asignatura__nombre'
    )

    filter_docente_id = request.GET.get('docente')
    filter_periodo_id = request.GET.get('periodo')

    if filter_docente_id:
        evaluaciones_query = evaluaciones_query.filter(docente__id=filter_docente_id)
    if filter_periodo_id:
        evaluaciones_query = evaluaciones_query.filter(periodo__id=filter_periodo_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Evaluacion Academica"

    # --- Configuración de Estilos ---
    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul oscuro para encabezados principales
    sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Azul claro para sub-encabezados
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    left_aligned_text = Alignment(horizontal="left", vertical="center")
    top_aligned_text = Alignment(horizontal="left", vertical="top", wrapText=True) 

    # --- DEFINICIÓN DE HEADERS (AÑADIENDO PUNTAJES Y JUICIOS) ---
    headers = [
        "Docente", "Cédula Docente", "Asignatura", "Carrera", "Semestre", "Período Académico", 
        "Fecha de Inicio Período", "Fecha de Fin Período", "Fue Evaluada", "Fecha de Evaluación", 
        "Docente Evaluador", 
        "Puntaje Acompañamiento", "Juicio Acompañamiento",
        "Puntaje Autoevaluación", "Juicio Autoevaluación",
        "Puntaje Evaluación Estudiante", "Juicio Evaluación Estudiante",
        "Juicio de Valor General" 
    ]

    current_row = 1 

    # --- INFORMACIÓN DE LA UNIVERSIDAD ---
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "UNIVERSIDAD NACIONAL EXPERIMENTAL POLITÉCNICA DE LA FUERZA ARMADA NACIONAL BOLIVARIANA"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    ws[f'A{current_row}'].alignment = center_aligned_text
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "VICERRECTORADO DE ASUNTOS ACADÉMICOS - UNIDAD DE PROGRAMACIÓN ACADÉMICA" 
    ws[f'A{current_row}'].font = Font(size=12, bold=True)
    ws[f'A{current_row}'].alignment = center_aligned_text
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "Núcleo: Falcón - Extensión: Punto Fijo"
    ws[f'A{current_row}'].font = Font(size=10, bold=True)
    ws[f'A{current_row}'].alignment = center_aligned_text
    current_row += 1

    current_row += 1 # Fila vacía para separación

    # Título del Reporte
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers)) 
    ws[f'A{current_row}'] = "REPORTE DE EVALUACIONES ACADÉMICAS" 
    ws[f'A{current_row}'].font = Font(size=16, bold=True, color="000000")
    ws[f'A{current_row}'].alignment = center_aligned_text
    ws.row_dimensions[current_row].height = 30 
    current_row += 1

    # Fecha del Reporte y Filtros Aplicados
    current_row += 1 
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=int(len(headers)/2)) 
    ws[f'A{current_row}'] = f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    filtro_info_list = []
    if filter_docente_id:
        try:
            docente_filtro = Docente.objects.get(id=filter_docente_id)
            filtro_info_list.append(f"Docente: {docente_filtro.nombre}") 
        except Docente.DoesNotExist:
            pass
    if filter_periodo_id:
        try:
            periodo_filtro = Periodo.objects.get(id=filter_periodo_id)
            filtro_info_list.append(f"Período: {periodo_filtro.nombre}")
        except Periodo.DoesNotExist:
            pass

    if filtro_info_list:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=int(len(headers)/2))
        ws[f'A{current_row}'] = "Filtros Aplicados:"
        ws[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        for f_text in filtro_info_list:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=int(len(headers)/2))
            ws[f'A{current_row}'] = f_text
            current_row += 1
    
    current_row += 2 

    # --- Sección de Estadísticas ---
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws[f'A{current_row}'] = "Resumen Estadístico"
    ws[f'A{current_row}'].font = Font(size=14, bold=True, color="000000")
    ws[f'A{current_row}'].alignment = left_aligned_text
    current_row += 1

    total_evaluaciones = evaluaciones_query.count()
    evaluaciones_realizadas = evaluaciones_query.filter(fue_evaluada=True).count()
    evaluaciones_pendientes = evaluaciones_query.filter(fue_evaluada=False).count()

    ws.append(["Total de Evaluaciones:", total_evaluaciones])
    ws.append(["Evaluaciones Realizadas:", evaluaciones_realizadas])
    ws.append(["Evaluaciones Pendientes:", evaluaciones_pendientes])
    
    # NUEVO: Estadísticas de puntajes y juicios (solo si hay datos)
    if evaluaciones_realizadas > 0:
        # Promedios de puntajes
        total_score_acompanamiento = sum(e.score_acompanamiento or 0 for e in evaluaciones_query if e.fue_evaluada and e.score_acompanamiento is not None)
        total_autoevaluacion_score = sum(e.autoevaluacion_score or 0 for e in evaluaciones_query if e.fue_evaluada and e.autoevaluacion_score is not None)
        total_evaluacion_estudiante_score = sum(e.evaluacion_estudiante_score or 0 for e in evaluaciones_query if e.fue_evaluada and e.evaluacion_estudiante_score is not None)

        count_acompanamiento = sum(1 for e in evaluaciones_query if e.fue_evaluada and e.score_acompanamiento is not None)
        count_autoevaluacion = sum(1 for e in evaluaciones_query if e.fue_evaluada and e.autoevaluacion_score is not None)
        count_evaluacion_estudiante = sum(1 for e in evaluaciones_query if e.fue_evaluada and e.evaluacion_estudiante_score is not None)

        avg_acompanamiento = (total_score_acompanamiento / count_acompanamiento) if count_acompanamiento > 0 else "N/A"
        avg_autoevaluacion = (total_autoevaluacion_score / count_autoevaluacion) if count_autoevaluacion > 0 else "N/A"
        avg_evaluacion_estudiante = (total_evaluacion_estudiante_score / count_evaluacion_estudiante) if count_evaluacion_estudiante > 0 else "N/A"
        
        ws.append(["Promedio Puntaje Acompañamiento:", f"{avg_acompanamiento:.2f}" if avg_acompanamiento != "N/A" else "N/A"])
        ws.append(["Promedio Puntaje Autoevaluación:", f"{avg_autoevaluacion:.2f}" if avg_autoevaluacion != "N/A" else "N/A"])
        ws.append(["Promedio Puntaje Evaluación Estudiante:", f"{avg_evaluacion_estudiante:.2f}" if avg_evaluacion_estudiante != "N/A" else "N/A"])

    current_row = ws.max_row + 1 
    current_row += 2 

    # --- Encabezados de la Tabla Principal de Datos ---
    ws.append(headers) 

    # Aplicar estilo a los encabezados
    header_row_index = current_row 
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_index, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_aligned_text
    current_row += 1 

    # --- Datos de la Tabla Principal (AÑADIENDO PUNTAJES Y JUICIOS) ---
    for evaluacion in evaluaciones_query:
        row_data = [
            evaluacion.docente.nombre if evaluacion.docente else 'N/A', 
            evaluacion.docente.cedula if evaluacion.docente else 'N/A',
            evaluacion.asignatura.nombre if evaluacion.asignatura else 'N/A',
            evaluacion.asignatura.carrera.nombre if evaluacion.asignatura and evaluacion.asignatura.carrera else 'N/A',
            evaluacion.asignatura.semestre.nombre if evaluacion.asignatura and evaluacion.asignatura.semestre else 'N/A',
            evaluacion.periodo.nombre if evaluacion.periodo else 'N/A',
            evaluacion.periodo.fecha_inicio.strftime('%d/%m/%Y') if evaluacion.periodo and evaluacion.periodo.fecha_inicio else 'N/A',
            evaluacion.periodo.fecha_fin.strftime('%d/%m/%Y') if evaluacion.periodo and evaluacion.periodo.fecha_fin else 'N/A',
            "Sí" if evaluacion.fue_evaluada else "No",
            evaluacion.fecha_evaluacion.strftime('%d/%m/%Y') if evaluacion.fecha_evaluacion else 'N/A',
            evaluacion.docente_evaluador.nombre if evaluacion.docente_evaluador else 'N/A', 
            
            # NUEVOS CAMPOS: Puntajes y Juicios
            evaluacion.score_acompanamiento if evaluacion.score_acompanamiento is not None else 'N/A',
            evaluacion.juicio_acompanamiento,
            evaluacion.autoevaluacion_score if evaluacion.autoevaluacion_score is not None else 'N/A',
            evaluacion.juicio_autoevaluacion,
            evaluacion.evaluacion_estudiante_score if evaluacion.evaluacion_estudiante_score is not None else 'N/A',
            evaluacion.juicio_evaluacion_estudiante,

            evaluacion.juicio_valor if evaluacion.juicio_valor else 'N/A', 
        ]
        ws.append(row_data) 
        
        for col_num in range(1, len(row_data) + 1):
            ws.cell(row=ws.max_row, column=col_num).border = thin_border
            # Aplicar alineación top a campos de texto largo
            if (col_num - 1) < len(headers) and \
            headers[col_num - 1] in ["Juicio de Valor General", 
                                    "Juicio Acompañamiento", 
                                    "Juicio Autoevaluación", 
                                    "Juicio Evaluación Estudiante"]: 
                ws.cell(row=ws.max_row, column=col_num).alignment = top_aligned_text

    # --- Autoajustar ancho de columnas ---
    # La variable header_row_index ya está definida justo antes del bucle de datos principales.
    # El bucle de autoajuste debe empezar a iterar sobre las celdas desde la primera fila de datos,
    # que es `header_row_index + 1`.
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx) 
        
        # Iterar sobre las celdas de la columna, empezando desde la primera fila de datos
        for row_idx in range(header_row_index + 1, ws.max_row + 1): 
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell_value = str(cell.value)
                if '\n' in cell_value:
                    lines = cell_value.split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
                else:
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
            except:
                pass
        
        adjusted_width = (max_length + 2) 
        if adjusted_width < 10: adjusted_width = 10 
        
        # Límite específico para las columnas de Juicio/Observaciones
        if (col_idx - 1) < len(headers) and \
        headers[col_idx - 1] in ["Juicio de Valor General", 
                                "Juicio Acompañamiento", 
                                "Juicio Autoevaluación", 
                                "Juicio Evaluación Estudiante"]:
            if adjusted_width > 60:
                adjusted_width = 60 
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_evaluacion_academica.xlsx"' 
    wb.save(response) 
    return response

